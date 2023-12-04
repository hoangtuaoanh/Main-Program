"""hoang.huan@maruichi_30/10/2023"""
# ghép code file stock
def gr_stock():
    import openpyxl
    import pprint
    wb = openpyxl.load_workbook(filename= 'Stock.xlsx')
    sheet = wb['Sheet1']
    max = sheet.max_row
    for i in range(7,max+1):

        cellB = sheet.cell(row=i,column=3)
        cellH = sheet.cell(row=i,column=8)
        cellI = sheet.cell(row=i,column=9)
        cellA = format(cellB.value) + format(cellH.value)+ format(cellI.value)
        sheet.cell(row=i,column=1,value=cellA)

    wb.save('Stock.xlsx')
gr_stock()
# ghép code file PO
def gr_po():
    import openpyxl
    import pprint
    ws = openpyxl.load_workbook(filename= 'PO.xlsx')
    sheet1 = ws['CUS030']
    max1 = sheet1.max_row
    for i in range(5,max1-4):

        cellB1 = sheet1.cell(row=i,column=2)
        cellM1 = sheet1.cell(row=i,column=13)
        cellI1 = sheet1.cell(row=i,column=9)
        cellA1 = format(cellB1.value) + format(cellM1.value)+ format(cellI1.value)
        sheet1.cell(row=i,column=1,value=cellA1)

    ws.save('PO.xlsx')
gr_po()
# đọc dữ liệu master 23/11/2023

def item():
    from openpyxl import load_workbook
    wi = load_workbook(filename= 'Route Master.xlsx')
    sheet_ranges = wi['route']
    maxi = sheet_ranges.max_row
    f = open("router.txt","a")
    list = []
    for i in range(3 , maxi+1):
        list.append(sheet_ranges[f"C{i}"].value)
    f.write(format(list))
    f.close()
item()
# đọc dữ liệu master 23/11/2023
def f_po():
    from openpyxl import load_workbook
    wp = load_workbook(filename= 'PO.xlsx')
    sheet_ranges = wp['CUS030']
    f = open("router.txt","r")
    list = f.read()
    masp = sheet_ranges.max_row
    for i in range( 5, masp-4):
        if sheet_ranges[f"A{i}"].value in list:
            continue
        else:
            ex=open("Diff.txt","a")
            ex.write(sheet_ranges[f"A{i}"].value)
        
    pass    
def f_stock():
    from openpyxl import load_workbook
    we = load_workbook(filename= 'Stock.xlsx')
    sheet_ranges = we['Sheet1']
    f = open("router.txt","r")
    list = f.read()
    masx = sheet_ranges.max_row
    for i in range( 7, masx+1):
        if sheet_ranges[f"A{i}"].value in list:
            continue
        else:
            ex=open("Diff.txt","a")
            t=sheet_ranges[f"A{i}"].value
            ex.write(f"{t}\n")
    ex.close
f_stock()
f_po()