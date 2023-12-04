def cal():
    from openpyxl import load_workbook
    wi = load_workbook(filename= 'Route Master.xlsx')
    sheet_ranges = wi['route']
    maxi = sheet_ranges.max_row
    List1 = []
    for i in range(3 , maxi+1):
        List1.append(sheet_ranges[f"E{i}"].value)
    pass
    group = set(List1)
    we = load_workbook(filename= 'End_Cal.xlsx')
    L = we['End_Cal']
    i = 3
    for x in group:
        L[f"E{i}"].value = x
        #print(L[f"E{i}"].value)
        i+=1
    we.save('End_Cal.xlsx')
cal()
